import json
import os
import tempfile
import io
import secrets
import re
from datetime import datetime
from pathlib import Path

from flask import Blueprint, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from werkzeug.utils import secure_filename

from db import db_cursor
from api_utils import success, fail, require_current_user


question_bank_bp = Blueprint("question_bank_api", __name__)
EXPORT_DIR = Path(__file__).resolve().parent / "exports"


QUESTION_TYPE_LABELS = {
    "single_choice": "单选题",
    "multiple_choice": "多选题",
    "true_false": "判断题",
    "blank": "填空题",
    "short_answer": "简答题",
    "essay": "论述题",
}


def to_bank_summary(row, current_user_id):
    return {
        "id": row["id"],
        "name": row["name"],
        "description": row.get("description") or "",
        "subjectName": row.get("subject_name") or "未分类",
        "folderId": int(row.get("folder_id") or 0),
        "folderName": row.get("folder_name") or "",
        "visibility": row.get("visibility") or "private",
        "sourceType": row.get("source_type") or "manual",
        "totalCount": int(row.get("total_count") or 0),
        "practicedCount": int(row.get("practiced_count") or 0),
        "wrongCount": int(row.get("wrong_count") or 0),
        "canManage": row.get("owner_user_id") == current_user_id,
        "createdAt": str(row.get("created_at") or ""),
        "updatedAt": str(row.get("updated_at") or ""),
    }


def fetch_bank_summary(cursor, bank_id, user_id):
    cursor.execute(
        """
        SELECT
          qb.id,
          qb.owner_user_id,
          qb.subject_id,
          qb.folder_id,
          qb.name,
          qb.description,
          qb.visibility,
          qb.source_type,
          qb.created_at,
          qb.updated_at,
          s.name AS subject_name,
          qbf.name AS folder_name,
          COUNT(DISTINCT q.id) AS total_count,
          COUNT(DISTINCT pa_done.question_id) AS practiced_count,
          COUNT(DISTINCT pa_wrong.question_id) AS wrong_count
        FROM question_banks qb
        LEFT JOIN subjects s ON s.id = qb.subject_id
        LEFT JOIN question_bank_folders qbf ON qbf.id = qb.folder_id
        LEFT JOIN questions q
          ON q.question_bank_id = qb.id
         AND q.status = 'active'
        LEFT JOIN practice_answers pa_done
          ON pa_done.question_id = q.id
         AND pa_done.user_id = %s
        LEFT JOIN practice_answers pa_wrong
          ON pa_wrong.question_id = q.id
         AND pa_wrong.user_id = %s
         AND pa_wrong.is_correct = 0
        WHERE qb.id = %s
          AND (
            qb.owner_user_id = %s OR qb.visibility IN ('public', 'class')
            OR EXISTS (
              SELECT 1
              FROM question_bank_shares qbs
              JOIN question_bank_share_members qbsm ON qbsm.share_id=qbs.id
              WHERE qbs.question_bank_id=qb.id
                AND qbs.status='active'
                AND (qbs.expires_at IS NULL OR qbs.expires_at>NOW())
                AND qbsm.user_id=%s
            )
          )
        GROUP BY
          qb.id, qb.owner_user_id, qb.subject_id, qb.folder_id, qb.name, qb.description,
          qb.visibility, qb.source_type, qb.created_at, qb.updated_at, s.name, qbf.name
        LIMIT 1
        """,
        (user_id, user_id, bank_id, user_id, user_id),
    )
    return cursor.fetchone()


@question_bank_bp.get("/question-banks")
def list_question_banks():
    user, error_response = require_current_user()
    if error_response:
        return error_response

    with db_cursor() as cursor:
        cursor.execute(
            """
            SELECT
              qb.id,
              qb.owner_user_id,
              qb.subject_id,
              qb.folder_id,
              qb.name,
              qb.description,
              qb.visibility,
              qb.source_type,
              qb.created_at,
              qb.updated_at,
              s.name AS subject_name,
              qbf.name AS folder_name,
              COUNT(DISTINCT q.id) AS total_count,
              COUNT(DISTINCT pa_done.question_id) AS practiced_count,
              COUNT(DISTINCT pa_wrong.question_id) AS wrong_count
            FROM question_banks qb
            LEFT JOIN subjects s ON s.id = qb.subject_id
            LEFT JOIN question_bank_folders qbf ON qbf.id = qb.folder_id
            LEFT JOIN questions q
              ON q.question_bank_id = qb.id
             AND q.status = 'active'
            LEFT JOIN practice_answers pa_done
              ON pa_done.question_id = q.id
             AND pa_done.user_id = %s
            LEFT JOIN practice_answers pa_wrong
              ON pa_wrong.question_id = q.id
             AND pa_wrong.user_id = %s
             AND pa_wrong.is_correct = 0
            WHERE qb.owner_user_id = %s
               OR qb.visibility IN ('public', 'class')
               OR EXISTS (
                 SELECT 1
                 FROM question_bank_shares qbs
                 JOIN question_bank_share_members qbsm ON qbsm.share_id=qbs.id
                 WHERE qbs.question_bank_id=qb.id
                   AND qbs.status='active'
                   AND (qbs.expires_at IS NULL OR qbs.expires_at>NOW())
                   AND qbsm.user_id=%s
               )
            GROUP BY
              qb.id, qb.owner_user_id, qb.subject_id, qb.folder_id, qb.name, qb.description,
              qb.visibility, qb.source_type, qb.created_at, qb.updated_at, s.name, qbf.name
            ORDER BY qb.updated_at DESC, qb.id DESC
            """,
            (user["id"], user["id"], user["id"], user["id"]),
        )
        rows = cursor.fetchall()

    banks = [to_bank_summary(row, user["id"]) for row in rows]
    return success({"banks": banks}, "题库列表获取成功")


@question_bank_bp.post("/question-banks")
def create_question_bank_api():
    user, error_response = require_current_user()
    if error_response:
        return error_response

    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()
    description = (body.get("description") or "").strip()

    if len(name) == 0:
        return fail("题库名称不能为空")
    if len(name) > 128:
        return fail("题库名称不能超过 128 个字符")

    with db_cursor(commit=True) as cursor:
        cursor.execute(
            """
            INSERT INTO question_banks
              (owner_user_id, subject_id, name, description, visibility, source_type)
            VALUES
              (%s, NULL, %s, %s, 'private', 'manual')
            """,
            (user["id"], name, description),
        )
        bank_id = cursor.lastrowid
        row = fetch_bank_summary(cursor, bank_id, user["id"])

    return success({"bank": to_bank_summary(row, user["id"])}, "题库创建成功", 201)


@question_bank_bp.put("/question-banks/<int:bank_id>")
def update_question_bank_api(bank_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response

    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()

    if len(name) == 0:
        return fail("题库名称不能为空")
    if len(name) > 128:
        return fail("题库名称不能超过 128 个字符")

    with db_cursor(commit=True) as cursor:
        cursor.execute(
            """
            UPDATE question_banks
            SET name = %s
            WHERE id = %s AND owner_user_id = %s
            """,
            (name, bank_id, user["id"]),
        )

        if cursor.rowcount == 0:
            return fail("只能修改自己创建的题库", 403)

        row = fetch_bank_summary(cursor, bank_id, user["id"])

    return success({"bank": to_bank_summary(row, user["id"])}, "题库名称已修改")


@question_bank_bp.delete("/question-banks/<int:bank_id>")
def delete_question_bank_api(bank_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response

    with db_cursor(commit=True) as cursor:
        cursor.execute(
            """
            DELETE FROM question_banks
            WHERE id = %s AND owner_user_id = %s
            """,
            (bank_id, user["id"]),
        )

        if cursor.rowcount == 0:
            return fail("只能删除自己创建的题库", 403)

    return success(None, "题库已删除")


@question_bank_bp.get("/question-bank-folders")
def list_question_bank_folders():
    user, error_response = require_current_user()
    if error_response:
        return error_response
    with db_cursor() as cursor:
        cursor.execute(
            """
            SELECT qbf.id, qbf.name, qbf.sort_order, COUNT(qb.id) AS bank_count
            FROM question_bank_folders qbf
            LEFT JOIN question_banks qb
              ON qb.folder_id=qbf.id AND qb.owner_user_id=qbf.user_id
            WHERE qbf.user_id=%s
            GROUP BY qbf.id, qbf.name, qbf.sort_order
            ORDER BY qbf.sort_order, qbf.id
            """,
            (user["id"],),
        )
        folders = [
            {
                "id": int(row["id"]),
                "name": row["name"],
                "sortOrder": int(row.get("sort_order") or 0),
                "bankCount": int(row.get("bank_count") or 0),
            }
            for row in cursor.fetchall()
        ]
    return success({"folders": folders})


@question_bank_bp.post("/question-bank-folders")
def create_question_bank_folder():
    user, error_response = require_current_user()
    if error_response:
        return error_response
    name = str((request.get_json(silent=True) or {}).get("name") or "").strip()
    if not name or len(name) > 80:
        return fail("文件夹名称不能为空且不能超过80个字符")
    try:
        with db_cursor(commit=True) as cursor:
            cursor.execute(
                "INSERT INTO question_bank_folders(user_id,name) VALUES(%s,%s)",
                (user["id"], name),
            )
            folder_id = cursor.lastrowid
    except Exception as exc:
        if "Duplicate" in str(exc):
            return fail("已存在同名文件夹")
        raise
    return success({"id": folder_id, "name": name, "sortOrder": 0, "bankCount": 0}, "文件夹已创建", 201)


@question_bank_bp.put("/question-bank-folders/<int:folder_id>")
def rename_question_bank_folder(folder_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response
    name = str((request.get_json(silent=True) or {}).get("name") or "").strip()
    if not name or len(name) > 80:
        return fail("文件夹名称不能为空且不能超过80个字符")
    with db_cursor(commit=True) as cursor:
        cursor.execute(
            "UPDATE question_bank_folders SET name=%s WHERE id=%s AND user_id=%s",
            (name, folder_id, user["id"]),
        )
        if cursor.rowcount == 0:
            return fail("文件夹不存在", 404)
    return success(None, "文件夹名称已更新")


@question_bank_bp.delete("/question-bank-folders/<int:folder_id>")
def delete_question_bank_folder(folder_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response
    with db_cursor(commit=True) as cursor:
        cursor.execute(
            "DELETE FROM question_bank_folders WHERE id=%s AND user_id=%s",
            (folder_id, user["id"]),
        )
        if cursor.rowcount == 0:
            return fail("文件夹不存在", 404)
    return success(None, "文件夹已删除，题库已移回未分类")


@question_bank_bp.put("/question-banks/batch/move")
def batch_move_question_banks():
    user, error_response = require_current_user()
    if error_response:
        return error_response
    body = request.get_json(silent=True) or {}
    bank_ids = body.get("bank_ids") or []
    folder_id = int(body.get("folder_id") or 0)
    if not isinstance(bank_ids, list) or not bank_ids:
        return fail("请选择要移动的题库")
    normalized_ids = sorted({int(item) for item in bank_ids if int(item) > 0})
    with db_cursor(commit=True) as cursor:
        if folder_id > 0:
            cursor.execute(
                "SELECT id FROM question_bank_folders WHERE id=%s AND user_id=%s",
                (folder_id, user["id"]),
            )
            if not cursor.fetchone():
                return fail("目标文件夹不存在", 404)
        placeholders = ",".join(["%s"] * len(normalized_ids))
        cursor.execute(
            f"""
            UPDATE question_banks
            SET folder_id=%s
            WHERE owner_user_id=%s AND id IN ({placeholders})
            """,
            [folder_id or None, user["id"], *normalized_ids],
        )
        moved_count = cursor.rowcount
    return success({"movedCount": moved_count}, f"已移动 {moved_count} 个题库")


@question_bank_bp.post("/question-banks/<int:bank_id>/share")
def create_question_bank_share(bank_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response
    with db_cursor(commit=True) as cursor:
        cursor.execute(
            "SELECT id FROM question_banks WHERE id=%s AND owner_user_id=%s",
            (bank_id, user["id"]),
        )
        if not cursor.fetchone():
            return fail("只能分享自己创建的题库", 403)
        cursor.execute(
            """
            SELECT share_code FROM question_bank_shares
            WHERE question_bank_id=%s AND owner_user_id=%s AND status='active'
            ORDER BY id DESC LIMIT 1
            """,
            (bank_id, user["id"]),
        )
        existing = cursor.fetchone()
        if existing:
            share_code = existing["share_code"]
        else:
            share_code = secrets.token_hex(4).upper()
            cursor.execute(
                """
                INSERT INTO question_bank_shares(question_bank_id,owner_user_id,share_code)
                VALUES(%s,%s,%s)
                """,
                (bank_id, user["id"], share_code),
            )
    return success({"shareCode": share_code}, "分享码已生成")


@question_bank_bp.post("/question-bank-shares/claim")
def claim_question_bank_share():
    user, error_response = require_current_user()
    if error_response:
        return error_response
    share_code = str((request.get_json(silent=True) or {}).get("share_code") or "").strip().upper()
    if not share_code:
        return fail("请输入分享码")
    with db_cursor(commit=True) as cursor:
        cursor.execute(
            """
            SELECT qbs.id, qbs.owner_user_id, qb.name
            FROM question_bank_shares qbs
            JOIN question_banks qb ON qb.id=qbs.question_bank_id
            WHERE qbs.share_code=%s AND qbs.status='active'
              AND (qbs.expires_at IS NULL OR qbs.expires_at>NOW())
            LIMIT 1
            """,
            (share_code,),
        )
        share = cursor.fetchone()
        if not share:
            return fail("分享码不存在或已失效", 404)
        if int(share["owner_user_id"]) == int(user["id"]):
            return fail("这是你自己的题库")
        cursor.execute(
            "INSERT IGNORE INTO question_bank_share_members(share_id,user_id) VALUES(%s,%s)",
            (share["id"], user["id"]),
        )
    return success({"bankName": share["name"]}, "共享题库已加入")


@question_bank_bp.get("/question-banks/<int:bank_id>/export")
def export_question_bank(bank_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response
    with db_cursor() as cursor:
        cursor.execute(
            "SELECT id,name FROM question_banks WHERE id=%s AND owner_user_id=%s",
            (bank_id, user["id"]),
        )
        bank = cursor.fetchone()
        if not bank:
            return fail("只能导出自己创建的题库", 403)
        cursor.execute(
            """
            SELECT q.id,q.type,q.stem,q.score,q.analysis,q.knowledge_point
            FROM questions q
            WHERE q.question_bank_id=%s AND q.status='active'
            ORDER BY q.id
            """,
            (bank_id,),
        )
        questions = cursor.fetchall()
        question_ids = [row["id"] for row in questions]
        options_map = {}
        answers_map = {}
        if question_ids:
            placeholders = ",".join(["%s"] * len(question_ids))
            cursor.execute(
                f"""
                SELECT question_id,option_key,content
                FROM question_options
                WHERE question_id IN ({placeholders})
                ORDER BY question_id,sort_order,option_key
                """,
                question_ids,
            )
            for row in cursor.fetchall():
                options_map.setdefault(row["question_id"], []).append(row)
            cursor.execute(
                f"""
                SELECT question_id,GROUP_CONCAT(answer_text ORDER BY is_primary DESC,id SEPARATOR '；') AS answer_text
                FROM question_answers
                WHERE question_id IN ({placeholders})
                GROUP BY question_id
                """,
                question_ids,
            )
            answers_map = {row["question_id"]: row.get("answer_text") or "" for row in cursor.fetchall()}

    export_format = str(request.args.get("format") or "xlsx").strip().lower()
    if export_format not in ("xlsx", "pdf"):
        return fail("导出格式只能是 xlsx 或 pdf")

    filename = safe_export_filename(bank["name"]) or f"question-bank-{bank_id}"
    if export_format == "pdf":
        output = build_question_bank_pdf(bank, questions, options_map, answers_map)
        save_export_copy(output, filename, "pdf")
        return send_file(
            output,
            as_attachment=True,
            download_name=f"{filename}.pdf",
            mimetype="application/pdf",
        )

    workbook = Workbook()
    paper = workbook.active
    paper.title = "试题卷"
    answers = workbook.create_sheet("答案解析")
    header_fill = PatternFill("solid", fgColor="DCEBFF")
    for sheet in (paper, answers):
        sheet.merge_cells("A1:F1")
        sheet["A1"] = bank["name"]
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 65
        sheet.column_dimensions["D"].width = 12
        sheet.column_dimensions["E"].width = 24
        sheet.column_dimensions["F"].width = 50
    paper.append(["序号", "题型", "题目", "分值", "知识点", "选项"])
    answers.append(["序号", "题型", "题目", "答案", "知识点", "解析"])
    for sheet in (paper, answers):
        for cell in sheet[2]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
    for index, question in enumerate(questions, start=1):
        option_text = "\n".join(
            f"{item['option_key']}. {item['content']}"
            for item in options_map.get(question["id"], [])
        )
        paper.append([
            index,
            QUESTION_TYPE_LABELS.get(question["type"], question["type"]),
            question["stem"],
            float(question.get("score") or 0),
            question.get("knowledge_point") or "",
            option_text,
        ])
        answers.append([
            index,
            QUESTION_TYPE_LABELS.get(question["type"], question["type"]),
            question["stem"],
            answers_map.get(question["id"], ""),
            question.get("knowledge_point") or "",
            question.get("analysis") or "",
        ])
    for sheet in (paper, answers):
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    save_export_copy(output, filename, "xlsx")
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{filename}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def safe_export_filename(value):
    filename = re.sub(r'[\\/:*?"<>|]+', "_", str(value or "")).strip(" .")
    return filename[:100]


def save_export_copy(output, filename, extension):
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    export_path = EXPORT_DIR / f"{filename}-{timestamp}.{extension}"
    export_path.write_bytes(output.getvalue())
    output.seek(0)
    return export_path


def escape_pdf_text(value):
    return (
        str(value or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )


def build_question_bank_pdf(bank, questions, options_map, answers_map):
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title=bank["name"],
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ChineseTitle",
        parent=styles["Title"],
        fontName="STSong-Light",
        fontSize=18,
        leading=24,
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    body_style = ParagraphStyle(
        "ChineseBody",
        parent=styles["BodyText"],
        fontName="STSong-Light",
        fontSize=10.5,
        leading=17,
        spaceAfter=5,
    )
    meta_style = ParagraphStyle(
        "ChineseMeta",
        parent=body_style,
        textColor=colors.HexColor("#667085"),
        fontSize=9,
        leading=14,
    )
    answer_style = ParagraphStyle(
        "ChineseAnswer",
        parent=body_style,
        textColor=colors.HexColor("#175CD3"),
    )

    story = [
        Paragraph(escape_pdf_text(bank["name"]), title_style),
        Paragraph(f"题目数量：{len(questions)}", meta_style),
        Spacer(1, 5 * mm),
    ]
    for index, question in enumerate(questions, start=1):
        type_label = QUESTION_TYPE_LABELS.get(question["type"], question["type"])
        story.append(
            Paragraph(
                f"{index}. [{escape_pdf_text(type_label)}] {escape_pdf_text(question['stem'])}"
                f"（{float(question.get('score') or 0):g}分）",
                body_style,
            )
        )
        option_rows = options_map.get(question["id"], [])
        if option_rows:
            option_data = [
                [
                    Paragraph(escape_pdf_text(item["option_key"]), body_style),
                    Paragraph(escape_pdf_text(item["content"]), body_style),
                ]
                for item in option_rows
            ]
            option_table = Table(option_data, colWidths=[12 * mm, 145 * mm])
            option_table.setStyle(
                TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ])
            )
            story.append(option_table)
        if question.get("knowledge_point"):
            story.append(
                Paragraph(f"知识点：{escape_pdf_text(question['knowledge_point'])}", meta_style)
            )
        story.append(Spacer(1, 4 * mm))

    story.append(PageBreak())
    story.append(Paragraph("答案与解析", title_style))
    for index, question in enumerate(questions, start=1):
        story.append(
            Paragraph(
                f"{index}. 答案：{escape_pdf_text(answers_map.get(question['id'], ''))}",
                answer_style,
            )
        )
        analysis = question.get("analysis") or "暂无解析"
        story.append(Paragraph(f"解析：{escape_pdf_text(analysis)}", body_style))
        story.append(Spacer(1, 3 * mm))

    document.build(story)
    output.seek(0)
    return output


@question_bank_bp.get("/question-banks/<int:bank_id>")
def question_bank_detail_api(bank_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response

    with db_cursor() as cursor:
        bank_row = fetch_bank_summary(cursor, bank_id, user["id"])
        if not bank_row:
            return fail("题库不存在或无权访问", 404)

        cursor.execute(
            """
            SELECT COUNT(*) AS total_questions
            FROM questions
            WHERE question_bank_id = %s AND status = 'active'
            """,
            (bank_id,),
        )
        total_questions = int((cursor.fetchone() or {}).get("total_questions") or 0)

        cursor.execute(
            """
            SELECT COUNT(DISTINCT q.id) AS practiced_questions
            FROM questions q
            JOIN practice_answers pa ON pa.question_id = q.id
            WHERE q.question_bank_id = %s
              AND q.status = 'active'
              AND pa.user_id = %s
            """,
            (bank_id, user["id"]),
        )
        practiced_questions = int((cursor.fetchone() or {}).get("practiced_questions") or 0)

        cursor.execute(
            """
            SELECT COUNT(DISTINCT q.id) AS wrong_questions
            FROM questions q
            JOIN practice_answers pa ON pa.question_id = q.id
            WHERE q.question_bank_id = %s
              AND q.status = 'active'
              AND pa.user_id = %s
              AND pa.is_correct = 0
            """,
            (bank_id, user["id"]),
        )
        wrong_questions = int((cursor.fetchone() or {}).get("wrong_questions") or 0)

        cursor.execute(
            """
            SELECT COUNT(DISTINCT q.id) AS needs_review_questions
            FROM questions q
            JOIN practice_answers pa ON pa.question_id = q.id
            WHERE q.question_bank_id = %s
              AND q.status = 'active'
              AND pa.user_id = %s
              AND pa.review_status = 'needs_review'
            """,
            (bank_id, user["id"]),
        )
        needs_review_questions = int((cursor.fetchone() or {}).get("needs_review_questions") or 0)

        cursor.execute(
            """
            SELECT
              COUNT(pa.id) AS answer_count,
              COALESCE(SUM(CASE WHEN pa.is_correct = 1 THEN 1 ELSE 0 END), 0) AS correct_answer_count
            FROM practice_answers pa
            JOIN questions q ON q.id = pa.question_id
            WHERE q.question_bank_id = %s
              AND q.status = 'active'
              AND pa.user_id = %s
            """,
            (bank_id, user["id"]),
        )
        answer_row = cursor.fetchone() or {}
        answer_count = int(answer_row.get("answer_count") or 0)
        correct_answer_count = int(answer_row.get("correct_answer_count") or 0)
        accuracy_rate = round(correct_answer_count * 100 / answer_count, 1) if answer_count > 0 else 0

        cursor.execute(
            """
            SELECT type, COUNT(*) AS count_value
            FROM questions
            WHERE question_bank_id = %s AND status = 'active'
            GROUP BY type
            ORDER BY count_value DESC
            """,
            (bank_id,),
        )
        type_rows = cursor.fetchall()

        cursor.execute(
            """
            SELECT difficulty, COUNT(*) AS count_value
            FROM questions
            WHERE question_bank_id = %s AND status = 'active'
            GROUP BY difficulty
            ORDER BY difficulty
            """,
            (bank_id,),
        )
        difficulty_rows = cursor.fetchall()

        cursor.execute(
            """
            SELECT t.name, COUNT(*) AS count_value
            FROM question_tags qt
            JOIN tags t ON t.id = qt.tag_id
            JOIN questions q ON q.id = qt.question_id
            WHERE q.question_bank_id = %s
              AND q.status = 'active'
            GROUP BY t.id, t.name
            ORDER BY count_value DESC
            LIMIT 8
            """,
            (bank_id,),
        )
        tag_rows = cursor.fetchall()

        cursor.execute(
            """
            SELECT
              q.id,
              q.type,
              q.stem,
              q.difficulty,
              q.knowledge_point,
              GROUP_CONCAT(DISTINCT t.name ORDER BY t.name SEPARATOR ',') AS tag_names
            FROM questions q
            LEFT JOIN question_tags qt ON qt.question_id = q.id
            LEFT JOIN tags t ON t.id = qt.tag_id
            WHERE q.question_bank_id = %s
              AND q.status = 'active'
            GROUP BY q.id, q.type, q.stem, q.difficulty, q.knowledge_point
            ORDER BY q.created_at DESC, q.id DESC
            LIMIT 8
            """,
            (bank_id,),
        )
        question_rows = cursor.fetchall()

    def chart_ratio(value):
        if total_questions <= 0:
            return 0
        return round(int(value or 0) * 100 / total_questions, 1)

    type_distribution = [
        {
            "label": QUESTION_TYPE_LABELS.get(row["type"], row["type"]),
            "value": int(row["count_value"] or 0),
            "ratio": chart_ratio(row["count_value"]),
        }
        for row in type_rows
    ]

    difficulty_distribution = [
        {
            "label": f"难度 {row['difficulty']}",
            "value": int(row["count_value"] or 0),
            "ratio": chart_ratio(row["count_value"]),
        }
        for row in difficulty_rows
    ]

    tag_distribution = [
        {
            "label": row["name"],
            "value": int(row["count_value"] or 0),
            "ratio": chart_ratio(row["count_value"]),
        }
        for row in tag_rows
    ]

    recent_questions = []
    for row in question_rows:
        tag_names = row.get("tag_names") or ""
        recent_questions.append({
            "id": row["id"],
            "type": row["type"],
            "typeLabel": QUESTION_TYPE_LABELS.get(row["type"], row["type"]),
            "stem": row["stem"],
            "difficulty": int(row["difficulty"] or 3),
            "knowledgePoint": row.get("knowledge_point") or "",
            "tags": [item for item in tag_names.split(",") if item],
        })

    stats = {
        "totalQuestions": total_questions,
        "practicedQuestions": practiced_questions,
        "unpracticedQuestions": max(total_questions - practiced_questions, 0),
        "wrongQuestions": wrong_questions,
        "needsReviewQuestions": needs_review_questions,
        "answerCount": answer_count,
        "correctAnswerCount": correct_answer_count,
        "accuracyRate": accuracy_rate,
    }

    return success({
        "bank": to_bank_summary(bank_row, user["id"]),
        "stats": stats,
        "typeDistribution": type_distribution,
        "difficultyDistribution": difficulty_distribution,
        "tagDistribution": tag_distribution,
        "recentQuestions": recent_questions,
    }, "题库详情获取成功")

def parse_int(value, default_value):
    try:
        return int(value)
    except Exception:
        return default_value


def parse_bool(value):
    return str(value or "").lower() in ("1", "true", "yes", "on")


def build_in_placeholders(values):
    return ",".join(["%s"] * len(values))


def fetch_question_filters(cursor, bank_id):
    cursor.execute(
        """
        SELECT DISTINCT type
        FROM questions
        WHERE question_bank_id=%s AND status='active'
        ORDER BY type
        """,
        (bank_id,),
    )
    type_rows = cursor.fetchall()

    cursor.execute(
        """
        SELECT DISTINCT difficulty
        FROM questions
        WHERE question_bank_id=%s AND status='active'
        ORDER BY difficulty
        """,
        (bank_id,),
    )
    difficulty_rows = cursor.fetchall()

    cursor.execute(
        """
        SELECT DISTINCT knowledge_point
        FROM questions
        WHERE question_bank_id=%s
          AND status='active'
          AND knowledge_point IS NOT NULL
          AND knowledge_point <> ''
        ORDER BY knowledge_point
        LIMIT 50
        """,
        (bank_id,),
    )
    knowledge_rows = cursor.fetchall()

    return {
        "types": [
            {
                "value": row["type"],
                "label": QUESTION_TYPE_LABELS.get(row["type"], row["type"]),
            }
            for row in type_rows
        ],
        "difficulties": [
            int(row["difficulty"] or 0)
            for row in difficulty_rows
            if row.get("difficulty") is not None
        ],
        "knowledgePoints": [
            row["knowledge_point"]
            for row in knowledge_rows
            if row.get("knowledge_point")
        ],
    }


@question_bank_bp.get("/question-banks/<int:bank_id>/questions")
def question_list_api(bank_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response

    keyword = (request.args.get("keyword") or "").strip()
    question_type = (request.args.get("type") or "").strip()
    knowledge_point = (request.args.get("knowledgePoint") or "").strip()
    practice_status = (request.args.get("practiceStatus") or "all").strip()
    wrong_only = parse_bool(request.args.get("wrongOnly"))

    difficulty = parse_int(request.args.get("difficulty"), 0)
    page = max(parse_int(request.args.get("page"), 1), 1)
    page_size = parse_int(request.args.get("pageSize"), 10)
    page_size = min(max(page_size, 1), 50)
    offset = (page - 1) * page_size

    with db_cursor() as cursor:
        bank_row = fetch_bank_summary(cursor, bank_id, user["id"])
        if not bank_row:
            return fail("题库不存在或无权访问", 404)

        filters = fetch_question_filters(cursor, bank_id)

        where_sql = """
        FROM questions q
        LEFT JOIN practice_answers last_pa ON last_pa.id = (
          SELECT pa2.id
          FROM practice_answers pa2
          WHERE pa2.user_id = %s
            AND pa2.question_id = q.id
          ORDER BY pa2.answered_at DESC, pa2.id DESC
          LIMIT 1
        )
        WHERE q.question_bank_id = %s
          AND q.status = 'active'
        """
        params = [user["id"], bank_id]

        if keyword:
            where_sql += " AND (q.stem LIKE %s OR q.knowledge_point LIKE %s) "
            params.extend([f"%{keyword}%", f"%{keyword}%"])

        if question_type:
            where_sql += " AND q.type = %s "
            params.append(question_type)

        if difficulty > 0:
            where_sql += " AND q.difficulty = %s "
            params.append(difficulty)

        if knowledge_point:
            where_sql += " AND q.knowledge_point LIKE %s "
            params.append(f"%{knowledge_point}%")

        if practice_status == "practiced":
            where_sql += " AND last_pa.id IS NOT NULL "
        elif practice_status == "unpracticed":
            where_sql += " AND last_pa.id IS NULL "

        if wrong_only:
            where_sql += " AND last_pa.id IS NOT NULL AND last_pa.is_correct = 0 "

        cursor.execute(
            f"""
            SELECT COUNT(*) AS total
            {where_sql}
            """,
            params,
        )
        total = int((cursor.fetchone() or {}).get("total") or 0)
        page_count = max((total + page_size - 1) // page_size, 1)

        cursor.execute(
            f"""
            SELECT
              q.id,
              q.type,
              q.stem,
              q.analysis,
              q.difficulty,
              q.score,
              q.knowledge_point,
              CASE
                WHEN last_pa.id IS NULL THEN 'unpracticed'
                WHEN last_pa.is_correct = 1 THEN 'correct'
                ELSE 'wrong'
              END AS practice_status,
              CASE
                WHEN last_pa.id IS NULL THEN -1
                WHEN last_pa.is_correct = 1 THEN 1
                ELSE 0
              END AS latest_is_correct
            {where_sql}
            ORDER BY q.id ASC
            LIMIT %s OFFSET %s
            """,
            params + [page_size, offset],
        )
        question_rows = cursor.fetchall()

        question_ids = [row["id"] for row in question_rows]

        options_map = {}
        answers_map = {}
        tags_map = {}

        if question_ids:
            placeholders = build_in_placeholders(question_ids)

            cursor.execute(
                f"""
                SELECT question_id, option_key, content, is_correct, sort_order
                FROM question_options
                WHERE question_id IN ({placeholders})
                ORDER BY question_id, sort_order, option_key
                """,
                question_ids,
            )
            for row in cursor.fetchall():
                question_id = row["question_id"]
                options_map.setdefault(question_id, []).append({
                    "optionKey": row["option_key"],
                    "content": row["content"],
                    "isCorrect": int(row.get("is_correct") or 0),
                })

            cursor.execute(
                f"""
                SELECT question_id,
                       GROUP_CONCAT(answer_text ORDER BY is_primary DESC, id ASC SEPARATOR '\\n') AS answer_text
                FROM question_answers
                WHERE question_id IN ({placeholders})
                GROUP BY question_id
                """,
                question_ids,
            )
            for row in cursor.fetchall():
                answers_map[row["question_id"]] = row.get("answer_text") or ""

            cursor.execute(
                f"""
                SELECT qt.question_id, t.name
                FROM question_tags qt
                JOIN tags t ON t.id = qt.tag_id
                WHERE qt.question_id IN ({placeholders})
                ORDER BY t.name
                """,
                question_ids,
            )
            for row in cursor.fetchall():
                tags_map.setdefault(row["question_id"], []).append(row["name"])

    questions = []
    for index, row in enumerate(question_rows):
        question_id = row["id"]
        questions.append({
            "id": question_id,
            "seq": offset + index + 1,
            "type": row["type"],
            "typeLabel": QUESTION_TYPE_LABELS.get(row["type"], row["type"]),
            "stem": row["stem"] or "",
            "analysis": row.get("analysis") or "",
            "difficulty": int(row.get("difficulty") or 1),
            "score": float(row.get("score") or 0),
            "knowledgePoint": row.get("knowledge_point") or "",
            "practiceStatus": row.get("practice_status") or "unpracticed",
            "latestIsCorrect": int(row.get("latest_is_correct") or -1),
            "answerText": answers_map.get(question_id, ""),
            "options": options_map.get(question_id, []),
            "tags": tags_map.get(question_id, []),
        })

    return success({
        "bank": to_bank_summary(bank_row, user["id"]),
        "questions": questions,
        "filters": filters,
        "pagination": {
            "page": page,
            "pageSize": page_size,
            "total": total,
            "pageCount": page_count,
        },
    }, "题目列表获取成功")


def _owned_bank(cursor, bank_id, user_id):
    cursor.execute(
        "SELECT id,name,subject_id FROM question_banks WHERE id=%s AND owner_user_id=%s LIMIT 1",
        (bank_id, user_id),
    )
    return cursor.fetchone()


def _question_payload(body):
    question_type = str(body.get("type") or "").strip()
    if question_type not in QUESTION_TYPE_LABELS:
        raise ValueError("题型不正确")
    stem = str(body.get("stem") or "").strip()
    if not stem:
        raise ValueError("题干不能为空")
    difficulty = max(1, min(5, int(body.get("difficulty") or 3)))
    score = float(body.get("score") or 1)
    if score <= 0:
        raise ValueError("分值必须大于 0")
    status = str(body.get("status") or "active").strip()
    if status not in {"draft", "active"}:
        raise ValueError("题目状态不正确")
    options = body.get("options") or []
    if not isinstance(options, list):
        raise ValueError("选项格式不正确")
    answer_text = str(body.get("answer_text") or "").strip()
    if not answer_text:
        raise ValueError("答案不能为空")
    return {
        "type": question_type,
        "stem": stem,
        "analysis": str(body.get("analysis") or "").strip(),
        "difficulty": difficulty,
        "score": score,
        "knowledge_point": str(body.get("knowledge_point") or "").strip(),
        "status": status,
        "options": options,
        "answer_text": answer_text,
    }


def _replace_question_children(cursor, question_id, payload):
    cursor.execute("DELETE FROM question_options WHERE question_id=%s", (question_id,))
    for index, option in enumerate(payload["options"], start=1):
        if not isinstance(option, dict):
            continue
        option_key = str(option.get("option_key") or option.get("optionKey") or "").strip().upper()
        content = str(option.get("content") or "").strip()
        if option_key and content:
            cursor.execute(
                """
                INSERT INTO question_options(question_id,option_key,content,is_correct,sort_order)
                VALUES(%s,%s,%s,%s,%s)
                """,
                (
                    question_id,
                    option_key,
                    content,
                    1 if option_key in [item.strip().upper() for item in payload["answer_text"].split(",")] else 0,
                    index,
                ),
            )
    cursor.execute("DELETE FROM question_answers WHERE question_id=%s", (question_id,))
    cursor.execute(
        "INSERT INTO question_answers(question_id,answer_text,is_primary) VALUES(%s,%s,1)",
        (question_id, payload["answer_text"]),
    )


@question_bank_bp.get("/question-banks/<int:bank_id>/editor-questions")
def editor_question_list(bank_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response
    with db_cursor() as cursor:
        bank = _owned_bank(cursor, bank_id, user["id"])
        if not bank:
            return fail("只能编辑自己创建的题库", 403)
        cursor.execute(
            """
            SELECT id,type,stem,analysis,difficulty,score,knowledge_point,ai_generated,status,created_at
            FROM questions
            WHERE question_bank_id=%s AND status IN ('draft','active')
            ORDER BY status='draft' DESC,id DESC
            """,
            (bank_id,),
        )
        rows = cursor.fetchall()
        ids = [row["id"] for row in rows]
        options_map, answers_map = {}, {}
        if ids:
            placeholders = build_in_placeholders(ids)
            cursor.execute(
                f"SELECT question_id,option_key,content,is_correct FROM question_options WHERE question_id IN ({placeholders}) ORDER BY question_id,sort_order,id",
                ids,
            )
            for row in cursor.fetchall():
                options_map.setdefault(row["question_id"], []).append({
                    "optionKey": row["option_key"],
                    "content": row["content"],
                    "isCorrect": int(row.get("is_correct") or 0),
                })
            cursor.execute(
                f"SELECT question_id,GROUP_CONCAT(answer_text ORDER BY is_primary DESC,id SEPARATOR ',') answer_text FROM question_answers WHERE question_id IN ({placeholders}) GROUP BY question_id",
                ids,
            )
            answers_map = {row["question_id"]: row.get("answer_text") or "" for row in cursor.fetchall()}
    questions = [{
        "id": int(row["id"]),
        "type": row["type"],
        "typeLabel": QUESTION_TYPE_LABELS.get(row["type"], row["type"]),
        "stem": row["stem"] or "",
        "analysis": row.get("analysis") or "",
        "difficulty": int(row.get("difficulty") or 3),
        "score": float(row.get("score") or 1),
        "knowledgePoint": row.get("knowledge_point") or "",
        "answerText": answers_map.get(row["id"], ""),
        "aiGenerated": int(row.get("ai_generated") or 0),
        "status": row.get("status") or "draft",
        "options": options_map.get(row["id"], []),
    } for row in rows]
    return success({"bankId": bank_id, "bankName": bank["name"], "questions": questions})


@question_bank_bp.post("/question-banks/<int:bank_id>/questions")
def create_question_api(bank_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response
    try:
        payload = _question_payload(request.get_json(silent=True) or {})
    except (ValueError, TypeError) as exc:
        return fail(str(exc))
    with db_cursor(commit=True) as cursor:
        bank = _owned_bank(cursor, bank_id, user["id"])
        if not bank:
            return fail("只能编辑自己创建的题库", 403)
        cursor.execute(
            """
            INSERT INTO questions(question_bank_id,subject_id,type,stem,analysis,difficulty,score,knowledge_point,ai_generated,status)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,0,%s)
            """,
            (bank_id, bank.get("subject_id"), payload["type"], payload["stem"], payload["analysis"],
             payload["difficulty"], payload["score"], payload["knowledge_point"], payload["status"]),
        )
        question_id = cursor.lastrowid
        _replace_question_children(cursor, question_id, payload)
    return success({"questionId": question_id}, "题目已创建", 201)


@question_bank_bp.put("/questions/<int:question_id>")
def update_question_api(question_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response
    try:
        payload = _question_payload(request.get_json(silent=True) or {})
    except (ValueError, TypeError) as exc:
        return fail(str(exc))
    with db_cursor(commit=True) as cursor:
        cursor.execute(
            """
            SELECT q.id FROM questions q JOIN question_banks qb ON qb.id=q.question_bank_id
            WHERE q.id=%s AND qb.owner_user_id=%s
            """,
            (question_id, user["id"]),
        )
        if not cursor.fetchone():
            return fail("题目不存在或无权编辑", 404)
        cursor.execute(
            """
            UPDATE questions SET type=%s,stem=%s,analysis=%s,difficulty=%s,score=%s,
              knowledge_point=%s,status=%s WHERE id=%s
            """,
            (payload["type"], payload["stem"], payload["analysis"], payload["difficulty"],
             payload["score"], payload["knowledge_point"], payload["status"], question_id),
        )
        _replace_question_children(cursor, question_id, payload)
    return success(None, "题目已更新")


@question_bank_bp.delete("/questions/<int:question_id>")
def delete_question_api(question_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response
    with db_cursor(commit=True) as cursor:
        cursor.execute(
            """
            DELETE q FROM questions q JOIN question_banks qb ON qb.id=q.question_bank_id
            WHERE q.id=%s AND qb.owner_user_id=%s
            """,
            (question_id, user["id"]),
        )
        if cursor.rowcount == 0:
            return fail("题目不存在或无权删除", 404)
    return success(None, "题目已删除")


@question_bank_bp.put("/question-banks/<int:bank_id>/questions/batch")
def batch_update_questions(bank_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response
    body = request.get_json(silent=True) or {}
    ids = body.get("question_ids") or []
    if not isinstance(ids, list) or not ids:
        return fail("请选择题目")
    question_ids = sorted({int(item) for item in ids if int(item) > 0})
    action = str(body.get("action") or "").strip()
    with db_cursor(commit=True) as cursor:
        if not _owned_bank(cursor, bank_id, user["id"]):
            return fail("只能编辑自己创建的题库", 403)
        placeholders = build_in_placeholders(question_ids)
        params = [bank_id, *question_ids]
        if action == "publish":
            cursor.execute(
                f"UPDATE questions SET status='active' WHERE question_bank_id=%s AND id IN ({placeholders})",
                params,
            )
        elif action == "score":
            score = float(body.get("score") or 0)
            if score <= 0:
                return fail("分值必须大于 0")
            cursor.execute(
                f"UPDATE questions SET score=%s WHERE question_bank_id=%s AND id IN ({placeholders})",
                [score, *params],
            )
        elif action == "delete":
            cursor.execute(
                f"DELETE FROM questions WHERE question_bank_id=%s AND id IN ({placeholders})",
                params,
            )
        else:
            return fail("不支持的批量操作")
        affected = cursor.rowcount
    return success({"affectedCount": affected}, f"已处理 {affected} 道题")

ALLOWED_EXCEL_EXTENSIONS = {".xlsx", ".xls"}


def normalize_header(value):
    return str(value or "").strip().lower()


def get_cell(row_data, *keys, default=""):
    for key in keys:
        value = row_data.get(key)
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    return default


def normalize_question_type(value):
    raw = str(value or "").strip().lower()

    mapping = {
        "single_choice": "single_choice",
        "单选": "single_choice",
        "单选题": "single_choice",
        "multiple_choice": "multiple_choice",
        "多选": "multiple_choice",
        "多选题": "multiple_choice",
        "true_false": "true_false",
        "判断": "true_false",
        "判断题": "true_false",
        "blank": "blank",
        "填空": "blank",
        "填空题": "blank",
        "short_answer": "short_answer",
        "简答": "short_answer",
        "简答题": "short_answer",
        "essay": "essay",
        "论述": "essay",
        "论述题": "essay",
    }

    return mapping.get(raw, "single_choice")


def parse_int_value(value, default_value):
    try:
        return int(value)
    except Exception:
        return default_value


def parse_float_value(value, default_value):
    try:
        return float(value)
    except Exception:
        return default_value


def split_tags(value):
    text = str(value or "").replace("，", ",")
    return [item.strip() for item in text.split(",") if item.strip()]


def create_or_get_tag(cursor, tag_name):
    cursor.execute(
        """
        INSERT INTO tags (name, category)
        VALUES (%s, 'knowledge')
        ON DUPLICATE KEY UPDATE id = LAST_INSERT_ID(id)
        """,
        (tag_name,),
    )
    return cursor.lastrowid


def parse_excel_rows(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if len(rows) <= 1:
        return []

    headers = [normalize_header(item) for item in rows[0]]
    parsed_rows = []

    for index, row in enumerate(rows[1:], start=2):
        row_data = {}
        has_value = False

        for col_index, header in enumerate(headers):
            if header == "":
                continue
            value = row[col_index] if col_index < len(row) else None
            if value is not None and str(value).strip() != "":
                has_value = True
            row_data[header] = value

        if has_value:
            parsed_rows.append({
                "row_number": index,
                "row_data": row_data,
            })

    return parsed_rows


@question_bank_bp.post("/question-banks/<int:bank_id>/import-excel")
def import_excel_questions_api(bank_id):
    user, error_response = require_current_user()
    if error_response:
        return error_response

    upload_file = request.files.get("file")
    if not upload_file:
        return fail("请上传 Excel 文件")

    original_file_name = secure_filename(upload_file.filename or "")
    _, ext = os.path.splitext(original_file_name.lower())

    if ext not in ALLOWED_EXCEL_EXTENSIONS:
        return fail("仅支持 .xlsx 或 .xls 文件")

    with db_cursor() as cursor:
        bank_row = fetch_bank_summary(cursor, bank_id, user["id"])
        if not bank_row:
            return fail("题库不存在或无权访问", 404)
        if bank_row["owner_user_id"] != user["id"]:
            return fail("只能向自己创建的题库导入题目", 403)

    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, f"ai_for_study_{user['id']}_{bank_id}_{original_file_name}")
    upload_file.save(temp_path)

    try:
        parsed_rows = parse_excel_rows(temp_path)
    except Exception as exc:
        return fail(f"Excel 解析失败：{exc}", 400)

    if len(parsed_rows) == 0:
        return fail("Excel 中没有可导入的题目数据")

    success_rows = 0
    failed_rows = 0
    import_id = 0
    error_messages = []

    with db_cursor(commit=True) as cursor:
        cursor.execute(
            """
            INSERT INTO question_imports
              (owner_user_id, question_bank_id, source_type, file_name, status,
               total_rows, success_rows, failed_rows, error_message)
            VALUES
              (%s, %s, 'excel', %s, 'processing', %s, 0, 0, NULL)
            """,
            (user["id"], bank_id, original_file_name, len(parsed_rows)),
        )
        import_id = cursor.lastrowid

        cursor.execute(
            "SELECT subject_id FROM question_banks WHERE id=%s",
            (bank_id,),
        )
        bank_subject = cursor.fetchone() or {}
        subject_id = bank_subject.get("subject_id")

        for item in parsed_rows:
            row_number = item["row_number"]
            row_data = item["row_data"]

            try:
                question_type = normalize_question_type(
                    get_cell(row_data, "type", "题型", "question_type", default="single_choice")
                )
                stem = get_cell(row_data, "stem", "题干", "question", "题目")
                answer = get_cell(row_data, "answer", "答案", "correct_answer")
                analysis = get_cell(row_data, "analysis", "解析", default="")
                knowledge_point = get_cell(row_data, "knowledge_point", "知识点", default="")
                difficulty = parse_int_value(get_cell(row_data, "difficulty", "难度", default="3"), 3)
                difficulty = max(1, min(5, difficulty))
                score = parse_float_value(get_cell(row_data, "score", "分值", default="1"), 1)

                if stem == "":
                    raise ValueError("题干不能为空")
                if answer == "":
                    raise ValueError("答案不能为空")

                cursor.execute(
                    """
                    INSERT INTO questions
                      (question_bank_id, subject_id, import_id, type, stem, analysis,
                       difficulty, score, knowledge_point, ai_generated, status, extra)
                    VALUES
                      (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0, 'active', %s)
                    """,
                    (
                        bank_id,
                        subject_id,
                        import_id,
                        question_type,
                        stem,
                        analysis,
                        difficulty,
                        score,
                        knowledge_point,
                        json.dumps({"excel_row": row_number}, ensure_ascii=False),
                    ),
                )
                question_id = cursor.lastrowid

                option_keys = ["A", "B", "C", "D", "E", "F"]
                answer_keys = [item.strip().upper() for item in answer.replace("，", ",").split(",")]

                if question_type in ("single_choice", "multiple_choice"):
                    for sort_order, option_key in enumerate(option_keys, start=1):
                        option_value = get_cell(
                            row_data,
                            f"option_{option_key.lower()}",
                            f"option_{option_key}",
                            f"选项{option_key}",
                            option_key,
                            default="",
                        )

                        if option_value != "":
                            cursor.execute(
                                """
                                INSERT INTO question_options
                                  (question_id, option_key, content, is_correct, sort_order)
                                VALUES
                                  (%s, %s, %s, %s, %s)
                                """,
                                (
                                    question_id,
                                    option_key,
                                    option_value,
                                    1 if option_key in answer_keys else 0,
                                    sort_order,
                                ),
                            )

                cursor.execute(
                    """
                    INSERT INTO question_answers
                      (question_id, answer_text, answer_json, is_primary)
                    VALUES
                      (%s, %s, %s, 1)
                    """,
                    (
                        question_id,
                        answer,
                        json.dumps({"answer": answer}, ensure_ascii=False),
                    ),
                )

                tag_text = get_cell(row_data, "tags", "标签", default="")
                all_tags = split_tags(tag_text)
                if knowledge_point and knowledge_point not in all_tags:
                    all_tags.append(knowledge_point)

                for tag_name in all_tags:
                    tag_id = create_or_get_tag(cursor, tag_name)
                    cursor.execute(
                        """
                        INSERT IGNORE INTO question_tags (question_id, tag_id)
                        VALUES (%s, %s)
                        """,
                        (question_id, tag_id),
                    )

                cursor.execute(
                    """
                    INSERT INTO question_import_rows
                      (import_id, row_number, raw_data, status, error_message, created_question_id)
                    VALUES
                      (%s, %s, %s, 'success', NULL, %s)
                    """,
                    (
                        import_id,
                        row_number,
                        json.dumps(row_data, ensure_ascii=False, default=str),
                        question_id,
                    ),
                )
                success_rows += 1

            except Exception as exc:
                failed_rows += 1
                error_message = f"第 {row_number} 行：{exc}"
                error_messages.append(error_message)

                cursor.execute(
                    """
                    INSERT INTO question_import_rows
                      (import_id, row_number, raw_data, status, error_message, created_question_id)
                    VALUES
                      (%s, %s, %s, 'failed', %s, NULL)
                    """,
                    (
                        import_id,
                        row_number,
                        json.dumps(row_data, ensure_ascii=False, default=str),
                        error_message,
                    ),
                )

        final_status = "success" if success_rows > 0 else "failed"
        joined_errors = "\n".join(error_messages[:5]) if error_messages else None

        cursor.execute(
            """
            UPDATE question_imports
            SET status=%s,
                success_rows=%s,
                failed_rows=%s,
                error_message=%s
            WHERE id=%s
            """,
            (final_status, success_rows, failed_rows, joined_errors, import_id),
        )

    if success_rows == 0:
        return fail(joined_errors or "没有题目导入成功", 400)

    return success({
        "importId": import_id,
        "totalRows": len(parsed_rows),
        "successRows": success_rows,
        "failedRows": failed_rows,
        "errorMessage": "\n".join(error_messages[:5]) if error_messages else "",
    }, "题目导入成功")
